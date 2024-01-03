import openpyxl


def getrowCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return (sheet.max_row)

def getcolmnCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return (sheet.max_column)

def readData(file,SheetName,rownum,coloumnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return sheet.cell(rownum,coloumnnum).value

def writeData(file,SheetName,rownum,coloumnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    sheet.cell(rownum,coloumnnum).value=data
    workbook.save(file)
